
import openpyxl



'''position = 0


def isExit_sheet_name(sheet_name):
    global position

    for pos, sheet in enumerate(wb.sheetnames):
        if sheet.lower() in sheet_name.lower():
            position = pos
            return True
    
    return False'''



   
def create_sheet(folder,game, league, performance, date):
    wb = None
    
    if folder == 'DRAW':
        wb = openpyxl.load_workbook(filename='BacktesteFevereiro.xlsx')

    else:
        if folder == 'UNDER':
            wb = openpyxl.load_workbook(filename='Backteste_underFevereiro.xlsx')

 
    #if isExit_sheet_name(sheet_name):
     #   print('this sheet is already exist')
      #  return

    



    sheet = wb.active


    sheet['A1'] = 'GAMES'
    sheet['B1'] = 'MARKET'
    sheet['C1'] = 'LEAGUE'
    sheet['D1'] = 'DATE'
    sheet['E1'] = 'PERFORMANCE'

    for row in range(2, 800):
        text = str(sheet.cell(row=row, column=1).value)
        
        if text == 'None':
            sheet.cell(row=row, column=1, value=game)
            sheet['C' + str(row)] = league
            sheet['D' + str(row)] = date
            sheet['E' + str(row)] = performance
            
            break
        continue
 

    #sheet['A1'] = 'GAMES'
    #sheet['B1'] = 'MARKET'

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['C'].width = 20
    
    if folder == 'DRAW':
        wb.save(filename='BacktesteFevereiro.xlsx')

    else:
        if folder == 'UNDER':
            wb.save(filename='Backteste_underFevereiro.xlsx')
    
